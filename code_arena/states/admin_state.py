"""Admin-side state: tests, students, submission review."""
from __future__ import annotations

import datetime as _dt
import io
import json
import openpyxl
import reflex as rx

from .. import db
from ..ai_grader import grade_submission
from ..config import settings
from ..logging_config import logger
from ..models import Role, SubmissionStatus, STATUS_LABELS
from ..tokens import make_export_token
from .auth_state import AuthState


def _fmt_epoch(value) -> str:
    try:
        return _dt.datetime.fromtimestamp(int(value)).strftime("%Y-%m-%d %H:%M")
    except (TypeError, ValueError, OverflowError):
        return "-"


class AdminState(AuthState):
    tests: list[dict[str, str]] = []
    students: list[dict[str, str]] = []
    submissions: list[dict[str, str]] = []

    status_msg: str = ""
    is_busy: bool = False

    @rx.var
    def has_tests(self) -> bool:
        return len(self.tests) > 0

    @rx.var
    def has_students(self) -> bool:
        return len(self.students) > 0

    @rx.var
    def has_submissions(self) -> bool:
        return len(self.submissions) > 0

    # --- create-test dialog fields --------------------------------------
    show_test_dialog: bool = False
    t_name: str = ""
    t_end: str = ""            # value from <input type=datetime-local>
    t_question: str = ""
    t_expected_input: str = ""
    t_expected_output: str = ""

    # --- add-student dialog fields --------------------------------------
    show_student_dialog: bool = False
    s_email: str = ""
    s_name: str = ""
    show_bulk_student_dialog: bool = False

    # --- review properties ----------------------------------------------
    show_review_dialog: bool = False
    review_sub_id: str = ""
    review_student: str = ""
    review_test: str = ""
    review_code: str = ""
    review_ai_grade: str = ""
    review_ai_feedback: str = ""
    review_final_grade: str = ""

    # --- drill-down state -----------------------------------------------
    selected_test_id: str = ""
    selected_sub_id: str = ""

    # --- review workspace state -----------------------------------------
    review_editor_files: dict[str, str] = {}
    review_folders: list[str] = []
    review_expanded_folders: list[str] = []
    review_active_file: str = "main.py"
    review_show_mobile_explorer: bool = False

    def set_t_name(self, value: str):
        self.t_name = value

    def set_t_end(self, value: str):
        self.t_end = value

    def set_t_question(self, value: str):
        self.t_question = value

    def set_t_expected_input(self, value: str):
        self.t_expected_input = value

    def set_t_expected_output(self, value: str):
        self.t_expected_output = value

    def set_s_email(self, value: str):
        self.s_email = value

    def set_s_name(self, value: str):
        self.s_name = value

    def set_review_final_grade(self, value: str):
        self.review_final_grade = value

    # ---------------------------------------------------------------
    # Loading
    # ---------------------------------------------------------------
    def load_dashboard(self):
        guard = self.require_admin()
        if guard is not None:
            return guard
        return AdminState.refresh_all

    def refresh_all(self):
        if not settings.appwrite_configured:
            self.status_msg = "Appwrite not configured — see .env.example."
            return
        self.is_busy = True
        yield
        try:
            self.tests = [
                {
                    "id": t["$id"],
                    "name": t.get("name", ""),
                    "end_at": _fmt_epoch(t.get("end_at")),
                    "question": t.get("question", ""),
                }
                for t in db.list_tests()
            ]
            self.students = [
                {
                    "id": u["$id"],
                    "name": u.get("name", ""),
                    "email": u.get("email", ""),
                }
                for u in db.list_users(role=Role.STUDENT)
            ]
            test_names = {t["id"]: t["name"] for t in self.tests}
            subs = []
            for s in db.list_all_submissions():
                subs.append(
                    {
                        "id": s["$id"],
                        "test_id": s.get("test_id", ""),
                        "test_name": test_names.get(s.get("test_id", ""), "?"),
                        "student": s.get("student_name", ""),
                        "status": s.get("status", ""),
                        "status_label": STATUS_LABELS.get(
                            s.get("status", ""), s.get("status", "")
                        ),
                        "ai_grade": s.get("ai_grade", ""),
                        "ai_feedback": s.get("ai_feedback", ""),
                        "final_grade": s.get("final_grade", ""),
                        "code": s.get("code", ""),
                        "files_json": s.get("files_json", "{}"),
                        "entry_file": s.get("entry_file", "main.py"),
                    }
                )
            self.submissions = subs
            self.status_msg = ""
        except (TypeError, ValueError, KeyError, RuntimeError) as exc:  # noqa: BLE001
            self.status_msg = f"Failed to load: {exc}"
        finally:
            self.is_busy = False

    # ---------------------------------------------------------------
    # Drill-down actions
    # ---------------------------------------------------------------
    def select_test(self, test_id: str):
        self.selected_test_id = test_id
        self.selected_sub_id = ""

    def clear_selected_test(self):
        self.selected_test_id = ""
        self.selected_sub_id = ""

    def select_sub(self, sub_id: str):
        self.selected_sub_id = sub_id
        match = next((s for s in self.submissions if s["id"] == sub_id), None)
        if not match:
            return
        self.review_sub_id = sub_id
        self.review_student = match["student"]
        self.review_test = match["test_name"]
        self.review_code = match["code"]
        self.review_ai_grade = match["ai_grade"]
        self.review_ai_feedback = match["ai_feedback"]
        self.review_final_grade = match["final_grade"] or match["ai_grade"]

        # Parse files_json
        try:
            files_dict = json.loads(match.get("files_json", "{}"))
        except Exception:
            files_dict = {"main.py": match["code"]}

        self.review_editor_files = files_dict
        self.review_active_file = match.get("entry_file", "main.py")
        if self.review_active_file not in self.review_editor_files and self.review_editor_files:
            self.review_active_file = list(self.review_editor_files.keys())[0]

        # Infer folders
        folders_set = set()
        for path in self.review_editor_files.keys():
            if "/" in path:
                parts = path.split("/")
                for i in range(1, len(parts)):
                    folders_set.add("/".join(parts[:i]))
        self.review_folders = list(folders_set)
        self.review_expanded_folders = list(folders_set)
        self.review_show_mobile_explorer = False

    def clear_selected_sub(self):
        self.selected_sub_id = ""

    @rx.var
    def test_submission_counts(self) -> dict[str, int]:
        counts = {}
        for s in self.submissions:
            tid = s.get("test_id")
            if tid:
                counts[tid] = counts.get(tid, 0) + 1
        return counts

    @rx.var
    def test_pending_counts(self) -> dict[str, int]:
        counts = {}
        for s in self.submissions:
            tid = s.get("test_id")
            if tid and s.get("status") in (SubmissionStatus.SUBMITTED, SubmissionStatus.AI_GRADED):
                counts[tid] = counts.get(tid, 0) + 1
        return counts

    @rx.var
    def filtered_submissions(self) -> list[dict]:
        return [s for s in self.submissions if s.get("test_id") == self.selected_test_id]

    @rx.var
    def selected_test_name(self) -> str:
        match = next((t for t in self.tests if t["id"] == self.selected_test_id), None)
        return match["name"] if match else ""

    @rx.var
    def has_filtered_submissions(self) -> bool:
        return len(self.filtered_submissions) > 0

    # ---------------------------------------------------------------
    # Review Workspace Logic (similar to StudentState)
    # ---------------------------------------------------------------
    @rx.var
    def review_active_content(self) -> str:
        return self.review_editor_files.get(self.review_active_file, "")

    @rx.var
    def review_file_tree_items(self) -> list[dict]:
        all_folders = set(self.review_folders)
        for path in self.review_editor_files.keys():
            if "/" in path:
                parts = path.split("/")
                for i in range(1, len(parts)):
                    all_folders.add("/".join(parts[:i]))

        nodes = []
        for f in all_folders:
            nodes.append({
                "name": f.split("/")[-1],
                "path": f,
                "kind": "folder",
                "depth": len(f.split("/")) - 1,
                "is_expanded": f in self.review_expanded_folders
            })
        for f in self.review_editor_files.keys():
            nodes.append({
                "name": f.split("/")[-1],
                "path": f,
                "kind": "file",
                "depth": len(f.split("/")) - 1 if "/" in f else 0,
                "is_expanded": False
            })

        # Sort: folders first, then files (hierarchically)
        def hierarchical_key(node):
            parts = node["path"].split("/")
            key_parts = []
            for i, part in enumerate(parts):
                is_last = (i == len(parts) - 1)
                is_folder = (node["kind"] == "folder") if is_last else True
                key_parts.append((0 if is_folder else 1, part))
            return key_parts

        nodes.sort(key=hierarchical_key)

        # Filter visibility
        visible_nodes = []
        for node in nodes:
            path = node["path"]
            parts = path.split("/")
            visible = True
            for i in range(1, len(parts)):
                ancestor = "/".join(parts[:i])
                if ancestor not in self.review_expanded_folders:
                    visible = False
                    break
            if visible:
                visible_nodes.append(node)

        return visible_nodes

    def select_review_path(self, path: str):
        if path in self.review_editor_files:
            self.review_active_file = path
            self.review_show_mobile_explorer = False
        else:
            self.toggle_review_folder(path)

    def toggle_review_folder(self, folder_path: str):
        if folder_path in self.review_expanded_folders:
            self.review_expanded_folders = [f for f in self.review_expanded_folders if f != folder_path]
        else:
            self.review_expanded_folders = self.review_expanded_folders + [folder_path]

    def toggle_review_mobile_explorer(self):
        self.review_show_mobile_explorer = not self.review_show_mobile_explorer

    def close_review_mobile_explorer(self):
        self.review_show_mobile_explorer = False

    # ---------------------------------------------------------------
    # Create test
    # ---------------------------------------------------------------
    def open_test_dialog(self):
        self.t_name = self.t_end = self.t_question = ""
        self.t_expected_input = self.t_expected_output = ""
        self.show_test_dialog = True

    def close_test_dialog(self):
        self.show_test_dialog = False

    def submit_test(self):
        if not self.t_name or not self.t_question or not self.t_end:
            self.status_msg = "Name, end date/time and question are required."
            logger.warning(f"Admin {self.user_email} test creation failed: missing fields.")
            return
        try:
            end_epoch = int(
                _dt.datetime.fromisoformat(self.t_end).timestamp()
            )
        except ValueError:
            self.status_msg = "Invalid end date/time."
            logger.warning(f"Admin {self.user_email} test creation failed: invalid end date format '{self.t_end}'.")
            return
        try:
            db.create_test(
                name=self.t_name,
                end_at=end_epoch,
                question=self.t_question,
                expected_input=self.t_expected_input,
                expected_output=self.t_expected_output,
                created_by=self.user_id,
            )
        except (TypeError, ValueError, RuntimeError) as exc:  # noqa: BLE001
            self.status_msg = f"Could not create test: {exc}"
            logger.error(f"Admin {self.user_email} test creation failed in DB: {exc}")
            return
        self.show_test_dialog = False
        self.status_msg = "Test created."
        logger.info(f"Admin {self.user_email} successfully created test '{self.t_name}'.")
        return AdminState.refresh_all

    def export_results(self):
        """Mint a fresh short-lived token and trigger the Excel download."""
        token = make_export_token()
        url = f"/api/export/results.xlsx?token={token}"
        # Opening the attachment URL downloads the file without navigating away.
        logger.info(f"Admin {self.user_email} exported test results sheet.")
        return rx.call_script(f"window.open('{url}', '_blank')")

    def delete_test(self, test_id: str):
        try:
            db.delete_test(test_id)
            self.status_msg = "Test deleted."
            logger.info(f"Admin {self.user_email} deleted test {test_id} successfully.")
        except (TypeError, ValueError, RuntimeError) as exc:  # noqa: BLE001
            self.status_msg = f"Delete failed: {exc}"
            logger.error(f"Admin {self.user_email} delete test {test_id} failed: {exc}")
        return AdminState.refresh_all

    # ---------------------------------------------------------------
    # Students
    # ---------------------------------------------------------------
    def open_student_dialog(self):
        self.s_email = self.s_name = ""
        self.show_student_dialog = True

    def close_student_dialog(self):
        self.show_student_dialog = False

    def open_bulk_student_dialog(self):
        self.show_bulk_student_dialog = True

    def close_bulk_student_dialog(self):
        self.show_bulk_student_dialog = False

    async def handle_excel_upload(self, files: list[rx.UploadFile]):
        logger.info(f"Admin {self.user_email} initiated bulk student Excel upload.")
        for file in files:
            try:
                upload_data = await file.read()
                logger.info(f"Read uploaded file '{file.filename}' (size: {file.size} bytes).")
                wb = openpyxl.load_workbook(io.BytesIO(upload_data), data_only=True)
                sheet = wb.active

                # Identify columns
                header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not header_row:
                    self.status_msg = "The uploaded Excel sheet is empty."
                    logger.warning(f"Excel upload failed: '{file.filename}' sheet is empty.")
                    continue

                name_idx = -1
                email_idx = -1
                for idx, val in enumerate(header_row):
                    if val:
                        val_lower = str(val).strip().lower()
                        if "name" in val_lower:
                            name_idx = idx
                        elif "email" in val_lower:
                            email_idx = idx

                # Fallbacks
                if name_idx == -1 or email_idx == -1:
                    name_idx = 0
                    email_idx = 1

                success_count = 0
                fail_count = 0

                # Read rows starting from row 2
                for row in list(sheet.iter_rows(min_row=2, values_only=True)):
                    if len(row) <= max(name_idx, email_idx):
                        continue
                    name_val = row[name_idx]
                    email_val = row[email_idx]

                    if not email_val:
                        continue

                    name = str(name_val).strip() if name_val else ""
                    email = str(email_val).strip().lower()

                    if not name:
                        name = email.split("@")[0]

                    try:
                        # Create user (email is passed, DB hashes it using Argon2)
                        db.create_user(
                            email=email,
                            password=email,
                            name=name,
                            role=Role.STUDENT
                        )
                        success_count += 1
                        logger.info(f"Created student account: '{email}' (Name: '{name}') via Excel import.")
                    except Exception as exc:
                        fail_count += 1
                        logger.warning(f"Failed to create student account '{email}' via Excel: {exc}")

                self.status_msg = f"Excel upload complete: created {success_count} student accounts ({fail_count} failed/skipped)."
                self.show_bulk_student_dialog = False
                logger.info(f"Excel upload complete for '{file.filename}': {success_count} created, {fail_count} failed/skipped.")

            except Exception as e:
                self.status_msg = f"Failed to process Excel file: {e}"
                logger.error(f"Failed to process Excel file '{file.filename}': {e}")

            finally:
                import os
                try:
                    if file.path and os.path.exists(file.path):
                        os.remove(file.path)
                        logger.info(f"Deleted local temporary file: '{file.path}'.")
                except Exception as exc:
                    logger.error(f"Failed to remove uploaded file {file.path}: {exc}")

        return AdminState.refresh_all

    def submit_student(self):
        email = self.s_email.strip().lower()
        if not email:
            self.status_msg = "Student email is required."
            logger.warning(f"Admin {self.user_email} student creation failed: email field empty.")
            return
        try:
            # Per spec: initial password equals the student's email address.
            db.create_user(
                email=email,
                password=email,
                name=self.s_name or email,
                role=Role.STUDENT,
            )
        except (TypeError, ValueError, RuntimeError) as exc:  # noqa: BLE001
            self.status_msg = f"Could not add student: {exc}"
            logger.error(f"Admin {self.user_email} failed to create student '{email}': {exc}")
            return
        self.show_student_dialog = False
        self.status_msg = f"Student {email} added (password = their email)."
        logger.info(f"Admin {self.user_email} added student '{email}' (Name: '{self.s_name or email}') successfully.")
        return AdminState.refresh_all

    def remove_student(self, student_id: str):
        try:
            db.delete_user(student_id)
            self.status_msg = "Student removed."
            logger.info(f"Admin {self.user_email} successfully removed student {student_id}.")
        except (TypeError, ValueError, RuntimeError) as exc:  # noqa: BLE001
            self.status_msg = f"Remove failed: {exc}"
            logger.error(f"Admin {self.user_email} failed to remove student {student_id}: {exc}")
        return AdminState.refresh_all

    # ---------------------------------------------------------------
    # AI grading trigger (grade all pending SUBMITTED)
    # ---------------------------------------------------------------
    def run_ai_grading(self):
        """Grade every submission still awaiting AI (status SUBMITTED)."""
        self.is_busy = True
        self.status_msg = "Running AI grading…"
        yield
        graded = 0
        try:
            for s in db.list_all_submissions():
                if s.get("status") != SubmissionStatus.SUBMITTED:
                    continue
                test = db.get_test(s.get("test_id", ""))
                if not test:
                    continue
                result = grade_submission(
                    question=test.get("question", ""),
                    expected_input=test.get("expected_input", ""),
                    expected_output=test.get("expected_output", ""),
                    code=s.get("code", ""),
                )
                db.set_ai_grade(s["$id"], result.grade, result.feedback)
                graded += 1
            self.status_msg = f"AI grading complete ({graded} submission(s))."
        except (TypeError, ValueError, KeyError, RuntimeError) as exc:  # noqa: BLE001
            self.status_msg = f"AI grading error: {exc}"
        finally:
            self.is_busy = False
        return AdminState.refresh_all

    # ---------------------------------------------------------------
    # Review one submission
    # ---------------------------------------------------------------
    def open_review(self, sub_id: str):
        self.select_sub(sub_id)

    def close_review(self):
        self.clear_selected_sub()

    def accept_grade(self):
        """Admin accepts (optionally edited) grade -> final."""
        if not self.review_final_grade:
            self.status_msg = "Enter a final grade before accepting."
            return
        try:
            db.set_final_grade(self.review_sub_id, self.review_final_grade)
            self.status_msg = "Final grade saved."
        except (TypeError, ValueError, RuntimeError) as exc:  # noqa: BLE001
            self.status_msg = f"Save failed: {exc}"
            return
        self.clear_selected_sub()
        return AdminState.refresh_all

    def redo_grade(self):
        """Admin rejects the AI grade -> re-queue for AI grading."""
        try:
            db.reset_to_ai_pending(self.review_sub_id)
            self.status_msg = "Sent back for AI re-grading."
        except (TypeError, ValueError, RuntimeError) as exc:  # noqa: BLE001
            self.status_msg = f"Action failed: {exc}"
            return
        self.clear_selected_sub()
        return AdminState.refresh_all
